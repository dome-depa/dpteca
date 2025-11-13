from django.core.management.base import BaseCommand
from django.db import transaction
from music.models import Artista
import openpyxl
import os
from pathlib import Path


class Command(BaseCommand):
    help = 'Importa artisti dal file Excel elencoArtisti.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='elencoArtisti.xlsx',
            help='Percorso del file Excel da importare (default: elencoArtisti.xlsx)'
        )
        parser.add_argument(
            '--skip-existing',
            action='store_true',
            help='Salta gli artisti che già esistono (basato sul nome)'
        )
        parser.add_argument(
            '--update-existing',
            action='store_true',
            help='Aggiorna gli artisti esistenti invece di crearne di nuovi'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        skip_existing = options['skip_existing']
        update_existing = options['update_existing']
        
        # Verifica se il file esiste
        if not os.path.exists(file_path):
            # Prova nella directory del progetto
            base_dir = Path(__file__).resolve().parent.parent.parent.parent
            file_path = base_dir / file_path
            if not file_path.exists():
                self.stdout.write(
                    self.style.ERROR(f'File non trovato: {file_path}')
                )
                return
        
        self.stdout.write(f'Lettura file: {file_path}')
        
        try:
            # Carica il workbook
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            self.stdout.write(f'Foglio: {ws.title}')
            self.stdout.write(f'Righe totali: {ws.max_row}')
            
            # Leggi i dati
            artisti_data = []
            header_skipped = False
            
            for row in ws.iter_rows(values_only=True):
                # Salta la riga header
                if not header_skipped:
                    header_skipped = True
                    continue
                
                # Prendi il primo valore (nome artista)
                nome_artista = row[0] if row else None
                
                if nome_artista:
                    # Pulisci il nome (rimuovi spazi extra, asterischi, etc.)
                    nome_artista = str(nome_artista).strip()
                    if nome_artista and nome_artista != 'None':
                        artisti_data.append(nome_artista)
            
            self.stdout.write(f'Artisti trovati nel file: {len(artisti_data)}')
            
            # Importa i dati
            created_count = 0
            updated_count = 0
            skipped_count = 0
            errors = []
            
            with transaction.atomic():
                for nome_artista in artisti_data:
                    try:
                        if update_existing:
                            # Aggiorna o crea
                            artista, created = Artista.objects.update_or_create(
                                nome_artista=nome_artista,
                                defaults={'nome_artista': nome_artista}
                            )
                            if created:
                                created_count += 1
                            else:
                                updated_count += 1
                        else:
                            # Controlla se esiste già
                            if skip_existing and Artista.objects.filter(nome_artista=nome_artista).exists():
                                skipped_count += 1
                                continue
                            
                            # Crea nuovo artista
                            artista, created = Artista.objects.get_or_create(
                                nome_artista=nome_artista
                            )
                            if created:
                                created_count += 1
                            else:
                                skipped_count += 1
                                
                    except Exception as e:
                        errors.append(f"Errore con '{nome_artista}': {str(e)}")
                        self.stdout.write(
                            self.style.WARNING(f"Errore con '{nome_artista}': {str(e)}")
                        )
            
            # Riepilogo
            self.stdout.write('')
            self.stdout.write(self.style.SUCCESS('=' * 50))
            self.stdout.write(self.style.SUCCESS('IMPORTAZIONE COMPLETATA'))
            self.stdout.write(self.style.SUCCESS('=' * 50))
            self.stdout.write(f'Artisti creati: {created_count}')
            if update_existing:
                self.stdout.write(f'Artisti aggiornati: {updated_count}')
            if skip_existing or not update_existing:
                self.stdout.write(f'Artisti saltati (già esistenti): {skipped_count}')
            if errors:
                self.stdout.write(self.style.ERROR(f'Errori: {len(errors)}'))
                for error in errors:
                    self.stdout.write(self.style.ERROR(f'  - {error}'))
            
            self.stdout.write('')
            self.stdout.write(f'Totale artisti nel database: {Artista.objects.count()}')
            
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Errore durante l\'importazione: {str(e)}')
            )
            import traceback
            self.stdout.write(traceback.format_exc())

